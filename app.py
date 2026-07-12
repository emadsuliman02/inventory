"""
نظام جرد وإدارة الأصول التقنية — سيرفر محلي على شبكة الجمعية
=================================================================
كيف يشتغل:
- يشغّله شخص واحد (مسؤول تقنية المعلومات) على جهاز واحد ثابت متصل بالشبكة.
- أي موظف على نفس الشبكة يفتح الرابط من متصفحه العادي (بدون تثبيت أي شيء).
- البيانات تُخزّن في ملفات JSON داخل مجلد data/ على نفس الجهاز — لا إنترنت، لا خدمات خارجية.

قبل التشغيل:
1) شغّل: pip install flask
2) شغّل: python app.py
3) أول تشغيل ينشئ حساب "administrator" مؤقت بكلمة مرور CHANGE_ME_ADMIN_PASSWORD —
   سجّل دخول فيه وغيّر كلمة المرور فورًا من تبويب "إدارة المستخدمين" داخل النظام.
4) اعرف عنوان IP لهذا الجهاز على الشبكة (راجع ملف التعليمات المرفق)
5) شارك الرابط: http://<عنوان-IP-لهذا-الجهاز>:5000 مع الموظفين على نفس الشبكة
"""

from flask import Flask, jsonify, request, session, send_from_directory, send_file
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook, load_workbook
from fpdf import FPDF
from datetime import date
import arabic_reshaper
from bidi.algorithm import get_display
import io
import json
import os
import re
import uuid
import secrets

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
STATIC_DIR = os.path.join(BASE_DIR, 'static')
SECRET_KEY_FILE = os.path.join(BASE_DIR, 'secret.key')

os.makedirs(DATA_DIR, exist_ok=True)

app = Flask(__name__, static_folder=STATIC_DIR)

# ------------------------------------------------------------------
# مفتاح تشفير الجلسات — يُنشأ تلقائيًا أول مرة ويبقى ثابتًا بعدها
# ------------------------------------------------------------------
if os.path.exists(SECRET_KEY_FILE):
    with open(SECRET_KEY_FILE, 'r') as f:
        app.secret_key = f.read().strip()
else:
    key = secrets.token_hex(32)
    with open(SECRET_KEY_FILE, 'w') as f:
        f.write(key)
    app.secret_key = key

# ------------------------------------------------------------------
# الحسابات — تُدار من تبويب "إدارة المستخدمين" داخل النظام نفسه
# (مسؤول النظام فقط). تُخزَّن في accounts.json محليًا على هذا الجهاز
# فقط (غير مرفوع لأي مستودع). أول تشغيل على جهاز جديد ينشئ حساب
# administrator مؤقت بكلمة مرور CHANGE_ME_ADMIN_PASSWORD لازم تغييرها فورًا.
# ------------------------------------------------------------------
ACCOUNTS_FILE = os.path.join(BASE_DIR, 'accounts.json')

# التبويبات القابلة للتخصيص لحسابات "مستخدم عادي" — لوحة المعلومات دايمًا
# متاحة للجميع، والقسم الآمن وإدارة المستخدمين دايمًا حصرية لمسؤول النظام.
ASSIGNABLE_PAGES = [
    'search', 'employees', 'devices', 'custody', 'stock', 'network', 'servers', 'security', 'doors',
    'submitrequest', 'requests', 'disposal',
]

DEFAULT_ACCOUNTS = {
    'administrator': {
        'password_hash': generate_password_hash('CHANGE_ME_ADMIN_PASSWORD'),
        'role': 'admin',
        'allowedPages': [],
    },
}


def load_accounts():
    if not os.path.exists(ACCOUNTS_FILE):
        save_accounts(DEFAULT_ACCOUNTS)
        return dict(DEFAULT_ACCOUNTS)
    with open(ACCOUNTS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_accounts(accounts):
    tmp_path = ACCOUNTS_FILE + '.tmp'
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(accounts, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, ACCOUNTS_FILE)


def allowed_pages_for(account):
    return ASSIGNABLE_PAGES if account['role'] == 'admin' else account.get('allowedPages', [])


TABLES = [
    'employees', 'devices', 'switches', 'routers', 'servers', 'security', 'doors', 'secure',
    'warehouses', 'stock_categories', 'stock_items', 'requests', 'disposal_requests',
]

# الجداول اللي يجوز طلب إتلاف قطعة منها
DISPOSABLE_TABLES = {'devices', 'servers', 'security', 'doors', 'switches', 'routers', 'stock_items'}
DISPOSAL_SOURCE_LABELS = {
    'devices': 'جهاز كمبيوتر', 'servers': 'سيرفر', 'security': 'جهاز أمان', 'doors': 'باب ذكي',
    'switches': 'سويتش', 'routers': 'راوتر', 'stock_items': 'قطعة مخزون',
}
ADMIN_ONLY_TABLES = {'secure'}

# جداول ما يُسمح بإضافة سجل جديد فيها مباشرة إلا لمسؤول النظام — أي إضافة أخرى لازم تمر بتبويب "تقديم طلب"
DIRECT_ADD_RESTRICTED_TABLES = {'devices', 'switches', 'routers', 'servers', 'security', 'doors'}

# الطلبات — الجدول الهدف اللي يتحول له الطلب عند القبول (سويتش/راوتر يتحدد حسب networkSubtype)
REQUEST_TARGET_TABLE = {
    'device': 'devices',
    'server': 'servers',
    'security': 'security',
    'employee': 'employees',
}
REQUEST_STATUSES = {'معلق', 'مقبول', 'مرجّع', 'مرفوض'}

# ------------------------------------------------------------------
# تصدير / استيراد Excel — ترتيب وتسمية الأعمدة لكل جدول
# ------------------------------------------------------------------
ASSET_SCHEMA = [
    ('name', 'اسم الجهاز'), ('location', 'الموقع'), ('floor', 'الدور'),
    ('custodian', 'بعهدة من'), ('model', 'الموديل'), ('specs', 'المواصفات'), ('status', 'الحالة'),
]

TABLE_SCHEMAS = {
    'employees': [
        ('name', 'الاسم'), ('empId', 'الرقم الوظيفي'), ('title', 'المسمى الوظيفي'),
        ('ext', 'رقم التحويلة'), ('email', 'البريد الإلكتروني'), ('emailPass', 'رمز البريد'),
        ('username', 'اسم المستخدم'), ('userPass', 'رمز اليوزر'), ('device', 'الجهاز المرتبط'),
        ('office', 'المكتب / القسم'), ('status', 'حالة الحساب'), ('notes', 'ملاحظات'),
    ],
    'devices': [
        ('code', 'اسم الجهاز'), ('type', 'النوع'), ('serial', 'الرقم التسلسلي'),
        ('notWorking', 'الجهاز لا يعمل'), ('user', 'الموظف المستلم للجهاز'), ('office', 'القسم'),
        ('ip', 'عنوان IP'), ('os', 'نظام التشغيل'), ('cpu', 'المعالج (CPU)'),
        ('mbModel', 'موديل اللوحة الأم'), ('mbSlots', 'سلوتات الرام المدعومة'), ('mbRamType', 'نوع الرام (DDR)'),
        ('ram', 'الرام (وصف)'), ('storage', 'الهارد ديسك (وصف)'), ('accessories', 'الملحقات'),
        ('licOffice', 'ترخيص Office'), ('licWindows', 'ترخيص Windows'), ('status', 'الحالة'), ('notes', 'ملاحظات'),
    ],
    'switches': [
        ('name', 'الاسم / الكود'), ('location', 'الموقع'), ('ports', 'عدد المنافذ'),
        ('ip', 'عنوان IP'), ('linked', 'مربوط بـ'), ('status', 'الحالة'),
    ],
    'routers': [
        ('name', 'الاسم / الكود'), ('location', 'الموقع'), ('ip', 'عنوان IP'),
        ('ssid', 'اسم الشبكة (SSID)'), ('status', 'الحالة'),
    ],
    'servers': ASSET_SCHEMA,
    'security': ASSET_SCHEMA,
    'doors': ASSET_SCHEMA,
    'secure': [
        ('linked', 'مرتبط بـ'), ('user', 'اسم المستخدم'), ('pass', 'كلمة المرور'), ('notes', 'ملاحظات'),
    ],
    'warehouses': [
        ('name', 'اسم المستودع'),
    ],
    'stock_categories': [
        ('name', 'اسم الفئة'), ('number', 'رقم الفئة'),
    ],
    'stock_items': [
        ('code', 'رمز القطعة'), ('name', 'اسم القطعة'), ('categoryName', 'الفئة'),
        ('categoryNumber', 'رقم الفئة'), ('warehouse', 'المستودع'),
        ('quantity', 'الكمية'), ('status', 'الحالة'),
    ],
}

BOOL_FIELDS = {'notWorking', 'licOffice', 'licWindows'}
ID_COLUMN_LABEL = 'المعرّف الداخلي (لا تعدله)'

# أسماء أعمدة بديلة شائعة (لملفات مصدرها خارجي، مو مصدّرة من النظام نفسه)
FIELD_ALIASES = {
    'employees': {
        'الاسم الكامل': 'name', 'اسم الموظف': 'name', 'اسم الموظف بالكامل': 'name',
        'الرقم الوظيفي للموظف': 'empId', 'رقم الموظف': 'empId',
        'الايميل': 'email', 'البريد': 'email', 'البريد الالكتروني': 'email',
        'اليوزر': 'username', 'اسم اليوزر': 'username',
        'الجهاز': 'device', 'القسم': 'office', 'المكتب': 'office',
        'الحالة': 'status',
    },
    'devices': {
        'المعرف': 'code', 'اسم الكمبيوتر': 'code', 'كود الجهاز': 'code',
        'المستخدم': 'user', 'الموظف': 'user',
    },
}


def normalize_header(text):
    text = (text or '').strip()
    text = re.sub(r'[ً-ْـ]', '', text)  # إزالة التشكيل والتطويل
    text = text.replace('أ', 'ا').replace('إ', 'ا').replace('آ', 'ا')
    text = text.replace('ة', 'ه').replace('ى', 'ي')
    return re.sub(r'\s+', ' ', text).strip().lower()


def label_lookup(table):
    lookup = {}
    for key, label in TABLE_SCHEMAS[table]:
        lookup[normalize_header(label)] = key
    for alias, key in FIELD_ALIASES.get(table, {}).items():
        lookup.setdefault(normalize_header(alias), key)
    return lookup


def record_to_row(table, record):
    row = []
    for key, _label in TABLE_SCHEMAS[table]:
        if table == 'devices' and key == 'mbModel':
            value = (record.get('motherboard') or {}).get('model', '')
        elif table == 'devices' and key == 'mbSlots':
            value = (record.get('motherboard') or {}).get('maxRamSlots', '')
        elif table == 'devices' and key == 'mbRamType':
            value = (record.get('motherboard') or {}).get('ramType', '')
        elif table == 'devices' and key == 'accessories':
            value = '، '.join(record.get('accessories') or [])
        elif key in BOOL_FIELDS:
            value = 'نعم' if record.get(key) else 'لا'
        else:
            value = record.get(key, '')
        row.append(value)
    return row


def row_to_record(table, row_dict, existing=None):
    record = dict(existing) if existing else {}
    motherboard = dict(record.get('motherboard') or {}) if table == 'devices' else None
    lookup = label_lookup(table)

    for header, raw in row_dict.items():
        key = lookup.get(normalize_header(header))
        if not key:
            continue
        value = '' if raw is None else str(raw).strip()

        if table == 'devices' and key == 'mbModel':
            motherboard['model'] = value
        elif table == 'devices' and key == 'mbSlots':
            motherboard['maxRamSlots'] = value
        elif table == 'devices' and key == 'mbRamType':
            motherboard['ramType'] = value
        elif table == 'devices' and key == 'accessories':
            record['accessories'] = [a.strip() for a in re.split('[,،]', value) if a.strip()]
        elif key in BOOL_FIELDS:
            record[key] = value in ('نعم', 'Yes', 'yes', 'true', '1')
        else:
            record[key] = value

    if table == 'devices':
        record['motherboard'] = motherboard
        record.setdefault('ramModules', (existing or {}).get('ramModules', []))
        record.setdefault('storageDisks', (existing or {}).get('storageDisks', []))
    if table in ('servers', 'security', 'doors'):
        record.setdefault('history', (existing or {}).get('history', []))

    return record


def data_path(table):
    return os.path.join(DATA_DIR, f'{table}.json')


def load_table(table):
    path = data_path(table)
    if not os.path.exists(path):
        return []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return []


def save_table(table, records):
    path = data_path(table)
    tmp_path = path + '.tmp'
    with open(tmp_path, 'w', encoding='utf-8') as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, path)  # كتابة آمنة تمنع تلف الملف عند انقطاع الكهرباء أثناء الحفظ


def current_role():
    return session.get('role')


def is_logged_in():
    return 'username' in session


# ------------------------------------------------------------------
# تسجيل الدخول / الخروج / التحقق من الجلسة
# ------------------------------------------------------------------
@app.route('/api/login', methods=['POST'])
def login():
    body = request.get_json(silent=True) or {}
    username = (body.get('username') or '').strip()
    password = body.get('password') or ''

    accounts = load_accounts()
    account = accounts.get(username)
    if not account or not check_password_hash(account['password_hash'], password):
        return jsonify({'error': 'بيانات الدخول غير صحيحة'}), 401

    session['username'] = username
    session['role'] = account['role']
    return jsonify({'username': username, 'role': account['role'], 'allowedPages': allowed_pages_for(account)})


@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})


@app.route('/api/session', methods=['GET'])
def get_session():
    if not is_logged_in():
        return jsonify({'loggedIn': False})
    accounts = load_accounts()
    account = accounts.get(session['username'])
    if not account:
        session.clear()
        return jsonify({'loggedIn': False})
    return jsonify({
        'loggedIn': True,
        'username': session['username'],
        'role': session['role'],
        'allowedPages': allowed_pages_for(account),
    })


# ------------------------------------------------------------------
# بيانات الجداول (موظفون / أجهزة / سويتشات / راوترات / القسم الآمن)
# ------------------------------------------------------------------
def require_login():
    if not is_logged_in():
        return jsonify({'error': 'الرجاء تسجيل الدخول أولاً'}), 401
    return None


def require_admin():
    if not is_logged_in():
        return jsonify({'error': 'الرجاء تسجيل الدخول أولاً'}), 401
    if current_role() != 'admin':
        return jsonify({'error': 'هذا القسم مخصص لمسؤول النظام فقط'}), 403
    return None


def has_page_access(page):
    if current_role() == 'admin':
        return True
    accounts = load_accounts()
    account = accounts.get(session.get('username')) or {}
    return page in (account.get('allowedPages') or [])


def require_page(page):
    if not is_logged_in():
        return jsonify({'error': 'الرجاء تسجيل الدخول أولاً'}), 401
    if not has_page_access(page):
        return jsonify({'error': 'ما عندك صلاحية الوصول لهذا القسم'}), 403
    return None


# ------------------------------------------------------------------
# إدارة المستخدمين (مسؤول النظام فقط)
# ------------------------------------------------------------------
@app.route('/api/users', methods=['GET'])
def list_users():
    err = require_admin()
    if err:
        return err
    accounts = load_accounts()
    return jsonify([
        {'username': u, 'role': a['role'], 'allowedPages': a.get('allowedPages', [])}
        for u, a in accounts.items()
    ])


@app.route('/api/users', methods=['POST'])
def upsert_user():
    err = require_admin()
    if err:
        return err

    body = request.get_json(silent=True) or {}
    username = (body.get('username') or '').strip()
    if not username:
        return jsonify({'error': 'اسم المستخدم إجباري'}), 400

    password = body.get('password') or ''
    role = 'admin' if body.get('role') == 'admin' else 'user'
    allowed_pages = [p for p in (body.get('allowedPages') or []) if p in ASSIGNABLE_PAGES]

    accounts = load_accounts()
    existing = accounts.get(username)

    if not password and not existing:
        return jsonify({'error': 'كلمة المرور إجبارية عند إنشاء حساب جديد'}), 400

    password_hash = generate_password_hash(password) if password else existing['password_hash']

    accounts[username] = {
        'password_hash': password_hash,
        'role': role,
        'allowedPages': allowed_pages,
    }
    save_accounts(accounts)

    if session.get('username') == username:
        session['role'] = role

    return jsonify({'username': username, 'role': role, 'allowedPages': allowed_pages})


@app.route('/api/users/<username>', methods=['DELETE'])
def delete_user(username):
    err = require_admin()
    if err:
        return err

    accounts = load_accounts()
    if username not in accounts:
        return jsonify({'error': 'المستخدم غير موجود'}), 404

    if username == session.get('username'):
        return jsonify({'error': 'لا يمكنك حذف حسابك الحالي وأنت مسجّل دخول فيه'}), 400

    remaining_admins = [u for u, a in accounts.items() if a['role'] == 'admin' and u != username]
    if accounts[username]['role'] == 'admin' and not remaining_admins:
        return jsonify({'error': 'لازم يبقى مسؤول نظام واحد على الأقل'}), 400

    del accounts[username]
    save_accounts(accounts)
    return jsonify({'success': True})


DEFAULT_WAREHOUSES = ['مستودع المحفوظات', 'مستودع وحدة التقنية']


def load_warehouses():
    warehouses = load_table('warehouses')
    if not warehouses:
        warehouses = [{'id': uuid.uuid4().hex, 'name': name} for name in DEFAULT_WAREHOUSES]
        save_table('warehouses', warehouses)
    return warehouses


@app.route('/api/<table>', methods=['GET'])
def get_table(table):
    if table not in TABLES:
        return jsonify({'error': 'جدول غير معروف'}), 404
    if table == 'disposal_requests':
        err = require_page('disposal')
        if err:
            return err
        return jsonify(load_table('disposal_requests'))
    err = require_admin() if table in ADMIN_ONLY_TABLES else require_login()
    if err:
        return err
    if table == 'warehouses':
        return jsonify(load_warehouses())
    if table == 'requests':
        records = load_table('requests')
        if not has_page_access('requests'):
            records = [r for r in records if r.get('submittedBy') == session.get('username')]
        return jsonify(records)
    return jsonify(load_table(table))


@app.route('/api/<table>', methods=['POST'])
def upsert_record(table):
    if table not in TABLES:
        return jsonify({'error': 'جدول غير معروف'}), 404
    if table == 'requests':
        return jsonify({'error': 'استخدم /api/requests/submit لتقديم الطلبات'}), 400
    if table == 'disposal_requests':
        return jsonify({'error': 'استخدم /api/disposal/request لتقديم طلبات الإتلاف'}), 400
    err = require_admin() if table in ADMIN_ONLY_TABLES else require_login()
    if err:
        return err

    record = request.get_json(silent=True) or {}
    records = load_table(table)

    if table == 'employees' and record.get('device'):
        owner = next((r for r in records if r.get('device') == record['device'] and r.get('id') != record.get('id')), None)
        if owner:
            return jsonify({'error': f"الجهاز {record['device']} بعهدة {owner.get('name')} بالفعل"}), 409

    existing_idx = None
    if record.get('id'):
        existing_idx = next((i for i, r in enumerate(records) if r.get('id') == record['id']), None)

    if existing_idx is None and table in DIRECT_ADD_RESTRICTED_TABLES and current_role() != 'admin':
        return jsonify({'error': 'الإضافة المباشرة غير متاحة — قدّم طلبًا من تبويب "تقديم طلب" وسيراجعه مسؤول النظام'}), 403

    if existing_idx is not None:
        records[existing_idx] = record
    elif record.get('id'):
        records.append(record)
    else:
        record['id'] = uuid.uuid4().hex
        records.append(record)

    save_table(table, records)
    return jsonify(record)


@app.route('/api/<table>/<record_id>', methods=['DELETE'])
def delete_record(table, record_id):
    if table not in TABLES:
        return jsonify({'error': 'جدول غير معروف'}), 404
    if table == 'disposal_requests':
        err = require_page('disposal')
    elif table in ADMIN_ONLY_TABLES or table == 'requests':
        err = require_admin()
    else:
        err = require_login()
    if err:
        return err

    records = load_table(table)
    records = [r for r in records if r.get('id') != record_id]
    save_table(table, records)
    return jsonify({'success': True})


# ------------------------------------------------------------------
# العهد — تسليم/استرجاع جهاز لموظف، مع مزامنة تبويبي الموظفين والأجهزة
# ------------------------------------------------------------------
@app.route('/api/custody/assign', methods=['POST'])
def assign_custody():
    err = require_login()
    if err:
        return err

    body = request.get_json(silent=True) or {}
    employee_id = body.get('employeeId')
    device_id = body.get('deviceId') or None

    employees = load_table('employees')
    devices = load_table('devices')

    employee = next((e for e in employees if e.get('id') == employee_id), None)
    if not employee:
        return jsonify({'error': 'الموظف غير موجود'}), 404

    new_device = None
    if device_id:
        new_device = next((d for d in devices if d.get('id') == device_id), None)
        if not new_device:
            return jsonify({'error': 'الجهاز غير موجود'}), 404
        holder = next(
            (e for e in employees if e.get('device') == new_device.get('code') and e.get('id') != employee_id),
            None,
        )
        if holder:
            return jsonify({'error': f"الجهاز {new_device['code']} بعهدة {holder.get('name')} بالفعل"}), 409

    today = date.today().isoformat()

    old_device_code = employee.get('device')
    if old_device_code and (not new_device or old_device_code != new_device.get('code')):
        old_device = next((d for d in devices if d.get('code') == old_device_code), None)
        if old_device:
            if old_device.get('user'):
                old_device.setdefault('history', []).append({'employee': old_device['user'], 'until': today})
            old_device['user'] = ''

    if new_device:
        previous_user = new_device.get('user')
        if previous_user and previous_user != employee.get('name'):
            new_device.setdefault('history', []).append({'employee': previous_user, 'until': today})
        new_device['user'] = employee.get('name')

    employee['device'] = new_device.get('code') if new_device else ''

    save_table('employees', employees)
    save_table('devices', devices)
    return jsonify({'success': True, 'employee': employee, 'device': new_device})


# ------------------------------------------------------------------
# المخزون — إدخال قطعة جديدة: يبحث عن الفئة بالاسم، يولّد رقم فئة
# جديد لو أول مرة، ويولّد رمز القطعة من (رقم الفئة-تسلسل)
# ------------------------------------------------------------------
@app.route('/api/stock/items', methods=['POST'])
def create_stock_item():
    err = require_login()
    if err:
        return err

    body = request.get_json(silent=True) or {}
    warehouse = (body.get('warehouse') or '').strip()
    category_name = (body.get('categoryName') or '').strip()
    name = (body.get('name') or '').strip()
    quantity = body.get('quantity')
    status = (body.get('status') or 'جديد').strip()

    if not warehouse:
        return jsonify({'error': 'اختر المستودع'}), 400
    if not category_name:
        return jsonify({'error': 'اكتب اسم الفئة'}), 400
    if not name:
        return jsonify({'error': 'اسم القطعة إجباري'}), 400

    categories = load_table('stock_categories')
    category = next((c for c in categories if c.get('name', '').strip() == category_name), None)
    if not category:
        next_number = max([c.get('number', 9999) for c in categories], default=9999) + 1
        category = {'id': uuid.uuid4().hex, 'name': category_name, 'number': next_number}
        categories.append(category)
        save_table('stock_categories', categories)

    items = load_table('stock_items')
    sequences = []
    for i in items:
        if i.get('categoryNumber') == category['number'] and '-' in (i.get('code') or ''):
            try:
                sequences.append(int(i['code'].rsplit('-', 1)[1]))
            except ValueError:
                pass
    next_seq = (max(sequences) + 1) if sequences else 1

    item = {
        'id': uuid.uuid4().hex,
        'code': f"{category['number']}-{next_seq:03d}",
        'name': name,
        'warehouse': warehouse,
        'categoryName': category['name'],
        'categoryNumber': category['number'],
        'quantity': quantity,
        'status': status,
    }
    items.append(item)
    save_table('stock_items', items)
    return jsonify(item)


# ------------------------------------------------------------------
# طلبات الإضافة — الموظف يقدّم طلب، المسؤول يقبل/يرجع/يرفض
# ------------------------------------------------------------------
@app.route('/api/requests/submit', methods=['POST'])
def submit_request():
    err = require_page('submitrequest')
    if err:
        return err

    body = request.get_json(silent=True) or {}
    request_type = body.get('requestType')
    if request_type not in REQUEST_TARGET_TABLE and request_type != 'network':
        return jsonify({'error': 'نوع الطلب غير معروف'}), 400

    network_subtype = body.get('networkSubtype') if request_type == 'network' else None
    if request_type == 'network' and network_subtype not in ('switch', 'router'):
        return jsonify({'error': 'حدد سويتش أو راوتر'}), 400

    fields = body.get('fields') or {}
    username = session['username']
    req_id = body.get('id')

    requests_list = load_table('requests')

    if req_id:
        existing = next((r for r in requests_list if r.get('id') == req_id), None)
        if not existing or existing.get('submittedBy') != username:
            return jsonify({'error': 'الطلب غير موجود'}), 404
        if existing.get('status') != 'مرجّع':
            return jsonify({'error': 'لا يمكن تعديل هذا الطلب في حالته الحالية'}), 400
        existing.update({
            'requestType': request_type,
            'networkSubtype': network_subtype,
            'fields': fields,
            'status': 'معلق',
            'adminNote': '',
            'submittedAt': date.today().isoformat(),
        })
        record = existing
    else:
        record = {
            'id': uuid.uuid4().hex,
            'requestType': request_type,
            'networkSubtype': network_subtype,
            'fields': fields,
            'submittedBy': username,
            'submittedAt': date.today().isoformat(),
            'status': 'معلق',
            'adminNote': '',
        }
        requests_list.append(record)

    save_table('requests', requests_list)
    return jsonify(record)


@app.route('/api/requests/<request_id>/decision', methods=['POST'])
def decide_request(request_id):
    err = require_page('requests')
    if err:
        return err

    body = request.get_json(silent=True) or {}
    decision = body.get('decision')
    note = (body.get('note') or '').strip()
    if decision not in ('accept', 'return', 'reject'):
        return jsonify({'error': 'إجراء غير معروف'}), 400

    requests_list = load_table('requests')
    req = next((r for r in requests_list if r.get('id') == request_id), None)
    if not req:
        return jsonify({'error': 'الطلب غير موجود'}), 404

    if decision == 'accept':
        target_table = REQUEST_TARGET_TABLE.get(req['requestType'])
        if req['requestType'] == 'network':
            target_table = 'switches' if req.get('networkSubtype') == 'switch' else 'routers'
        records = load_table(target_table)
        new_record = dict(req.get('fields') or {})
        new_record['id'] = uuid.uuid4().hex
        if target_table == 'devices':
            new_record.setdefault('status', 'يعمل')
        elif target_table == 'employees':
            new_record.setdefault('status', 'نشط')
        elif target_table in ('switches', 'routers'):
            new_record.setdefault('status', 'متصل')
        elif target_table == 'servers' or target_table == 'security':
            new_record.setdefault('status', 'متاح')
        records.append(new_record)
        save_table(target_table, records)
        req['status'] = 'مقبول'
        req['adminNote'] = note
    elif decision == 'return':
        req['status'] = 'مرجّع'
        req['adminNote'] = note
    else:
        req['status'] = 'مرفوض'
        req['adminNote'] = note

    save_table('requests', requests_list)
    return jsonify(req)


# ------------------------------------------------------------------
# طلبات الإتلاف — تسحب قطعة من أي جدول أصول، وموافقة المسؤول تحذفها
# نهائيًا من جدولها الأصلي وتخليها فقط بأرشيف طلبات الإتلاف
# ------------------------------------------------------------------
@app.route('/api/disposal/request', methods=['POST'])
def create_disposal_request():
    err = require_page('disposal')
    if err:
        return err

    body = request.get_json(silent=True) or {}
    source_table = body.get('sourceTable')
    source_id = body.get('sourceId')
    reason = (body.get('reason') or '').strip()

    if source_table not in DISPOSABLE_TABLES:
        return jsonify({'error': 'هذا النوع غير قابل لطلب الإتلاف'}), 400

    records = load_table(source_table)
    item = next((r for r in records if r.get('id') == source_id), None)
    if not item:
        return jsonify({'error': 'القطعة غير موجودة'}), 404

    disposal_list = load_table('disposal_requests')
    new_req = {
        'id': uuid.uuid4().hex,
        'sourceTable': source_table,
        'sourceId': source_id,
        'snapshot': item,
        'reason': reason,
        'requestedBy': session['username'],
        'requestedAt': date.today().isoformat(),
        'status': 'معلق',
        'decidedAt': None,
        'decidedBy': None,
        'decisionNote': '',
    }
    disposal_list.append(new_req)
    save_table('disposal_requests', disposal_list)
    return jsonify(new_req)


@app.route('/api/disposal/<request_id>/decision', methods=['POST'])
def decide_disposal(request_id):
    err = require_page('disposal')
    if err:
        return err

    body = request.get_json(silent=True) or {}
    decision = body.get('decision')
    note = (body.get('note') or '').strip()
    if decision not in ('approve', 'reject'):
        return jsonify({'error': 'إجراء غير معروف'}), 400

    disposal_list = load_table('disposal_requests')
    req = next((r for r in disposal_list if r.get('id') == request_id), None)
    if not req:
        return jsonify({'error': 'الطلب غير موجود'}), 404
    if req.get('status') != 'معلق':
        return jsonify({'error': 'تم اتخاذ قرار بهذا الطلب مسبقًا'}), 400

    if decision == 'approve':
        records = load_table(req['sourceTable'])
        records = [r for r in records if r.get('id') != req['sourceId']]
        save_table(req['sourceTable'], records)
        req['status'] = 'تمت الموافقة'
    else:
        req['status'] = 'تم الرفض'

    req['decidedAt'] = date.today().isoformat()
    req['decidedBy'] = session['username']
    req['decisionNote'] = note
    save_table('disposal_requests', disposal_list)
    return jsonify(req)


# ------------------------------------------------------------------
# تصدير PDF (بطاقة موظف / بطاقة جهاز / إيصال عهدة)
# ------------------------------------------------------------------
PDF_FONT_REGULAR = r'C:\Windows\Fonts\tahoma.ttf'
PDF_FONT_BOLD = r'C:\Windows\Fonts\tahomabd.ttf'
PDF_LOGO = os.path.join(STATIC_DIR, 'assets', 'icon.png')
PDF_ACCENT = (0xCE, 0x61, 0x28)
PDF_TEXT = (0x57, 0x56, 0x50)
PDF_MUTED = (0x91, 0x86, 0x7D)
PDF_LINE = (0xE7, 0xDC, 0xCB)


def rtl(text):
    text = '' if text is None else str(text)
    if not text:
        return ''
    return get_display(arabic_reshaper.reshape(text))


def new_pdf():
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.add_font('Body', '', PDF_FONT_REGULAR)
    pdf.add_font('Body', 'B', PDF_FONT_BOLD)
    pdf.add_page()
    pdf.set_margin(18)
    if os.path.exists(PDF_LOGO):
        pdf.image(PDF_LOGO, x=178, y=12, w=14)
    return pdf


def pdf_title(pdf, title, subtitle=None):
    pdf.set_font('Body', 'B', 20)
    pdf.set_text_color(*PDF_ACCENT)
    pdf.set_xy(18, 14)
    pdf.cell(155, 10, rtl(title), align='R')
    if subtitle:
        pdf.set_font('Body', '', 11)
        pdf.set_text_color(*PDF_MUTED)
        pdf.set_xy(18, 25)
        pdf.cell(155, 7, rtl(subtitle), align='R')
    pdf.set_draw_color(*PDF_LINE)
    pdf.line(18, 34, 192, 34)
    pdf.set_y(40)


def pdf_row(pdf, label, value):
    # reshape/bidi label and value separately (not the concatenated string) so a
    # multi-line value wraps in correct reading order and parentheses in labels
    # like "(CPU)" don't get mirrored by the value's own bidi context.
    pdf.set_font('Body', '', 12)
    pdf.set_text_color(*PDF_TEXT)
    value_text = str(value) if value not in (None, '') else '—'
    label_visual = rtl(label) + '  :'

    words = value_text.split(' ')
    lines = []
    current = ''
    for word in words:
        candidate = (current + ' ' + word).strip()
        probe_width = pdf.get_string_width(rtl(candidate) + '  ' + label_visual)
        if probe_width <= 174 or not current:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)

    for i, line in enumerate(lines):
        pdf.set_x(18)
        text_visual = (rtl(line) + '  ' + label_visual) if i == 0 else rtl(line)
        pdf.cell(174, 8, text_visual, align='R')
        pdf.ln(8)
    pdf.set_draw_color(*PDF_LINE)
    pdf.line(18, pdf.get_y(), 192, pdf.get_y())
    pdf.ln(2)


def pdf_paragraph(pdf, text, width=174, line_height=7):
    words = text.split(' ')
    lines = []
    current = ''
    for word in words:
        candidate = (current + ' ' + word).strip()
        if pdf.get_string_width(rtl(candidate)) <= width or not current:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    for line in lines:
        pdf.set_x(18)
        pdf.cell(width, line_height, rtl(line), align='R')
        pdf.ln(line_height)


def pdf_section(pdf, title):
    pdf.ln(3)
    pdf.set_font('Body', 'B', 13)
    pdf.set_text_color(*PDF_ACCENT)
    pdf.set_xy(18, pdf.get_y())
    pdf.cell(174, 9, rtl(title), align='R')
    pdf.ln(9)


def pdf_response(pdf, filename):
    buf = io.BytesIO(pdf.output())
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/pdf')


@app.route('/api/employees/<employee_id>/pdf', methods=['GET'])
def employee_pdf(employee_id):
    err = require_login()
    if err:
        return err
    employee = next((e for e in load_table('employees') if e.get('id') == employee_id), None)
    if not employee:
        return jsonify({'error': 'الموظف غير موجود'}), 404

    pdf = new_pdf()
    pdf_title(pdf, 'بطاقة بيانات الموظف', 'جمعية الإحسان للخدمات الاجتماعية')
    pdf_row(pdf, 'الاسم', employee.get('name'))
    pdf_row(pdf, 'الرقم الوظيفي', employee.get('empId'))
    pdf_row(pdf, 'المسمى الوظيفي', employee.get('title'))
    pdf_row(pdf, 'رقم التحويلة', employee.get('ext'))
    pdf_row(pdf, 'البريد الإلكتروني', employee.get('email'))
    pdf_row(pdf, 'اسم المستخدم', employee.get('username'))
    pdf_row(pdf, 'الجهاز المرتبط', employee.get('device'))
    pdf_row(pdf, 'المكتب / القسم', employee.get('office'))
    pdf_row(pdf, 'حالة الحساب', employee.get('status'))
    if employee.get('notes'):
        pdf_row(pdf, 'ملاحظات', employee.get('notes'))
    return pdf_response(pdf, f"employee-{employee.get('name') or employee_id}.pdf")


@app.route('/api/devices/<device_id>/pdf', methods=['GET'])
def device_pdf(device_id):
    err = require_login()
    if err:
        return err
    device = next((d for d in load_table('devices') if d.get('id') == device_id), None)
    if not device:
        return jsonify({'error': 'الجهاز غير موجود'}), 404

    mb = device.get('motherboard') or {}
    pdf = new_pdf()
    pdf_title(pdf, 'بطاقة بيانات الجهاز', 'جمعية الإحسان للخدمات الاجتماعية')
    pdf_row(pdf, 'اسم الجهاز', device.get('code'))
    pdf_row(pdf, 'النوع', device.get('type'))
    pdf_row(pdf, 'الرقم التسلسلي', device.get('serial'))
    pdf_row(pdf, 'الحالة', device.get('status'))
    pdf_row(pdf, 'بعهدة', device.get('user') or 'بدون مستخدم')
    pdf_row(pdf, 'القسم', device.get('office'))
    pdf_row(pdf, 'عنوان IP', device.get('ip'))
    pdf_row(pdf, 'نظام التشغيل', device.get('os'))
    pdf_row(pdf, 'المعالج (CPU)', device.get('cpu'))

    pdf_section(pdf, 'اللوحة الأم والذاكرة')
    pdf_row(pdf, 'موديل اللوحة الأم', mb.get('model'))
    pdf_row(pdf, 'نوع الرام (DDR)', mb.get('ramType'))
    pdf_row(pdf, 'الرام', device.get('ram'))
    pdf_row(pdf, 'الهارد ديسك', device.get('storage'))

    pdf_section(pdf, 'الملحقات والتراخيص')
    pdf_row(pdf, 'الملحقات', '، '.join(device.get('accessories') or []))
    pdf_row(pdf, 'ترخيص Office', 'فعالة' if device.get('licOffice') else 'غير فعالة')
    pdf_row(pdf, 'ترخيص Windows', 'فعالة' if device.get('licWindows') else 'غير فعالة')
    if device.get('notes'):
        pdf_row(pdf, 'ملاحظات', device.get('notes'))
    return pdf_response(pdf, f"device-{device.get('code') or device_id}.pdf")


@app.route('/api/custody/<employee_id>/pdf', methods=['GET'])
def custody_pdf(employee_id):
    err = require_login()
    if err:
        return err
    employee = next((e for e in load_table('employees') if e.get('id') == employee_id), None)
    if not employee:
        return jsonify({'error': 'الموظف غير موجود'}), 404
    device = next((d for d in load_table('devices') if d.get('code') == employee.get('device')), None)

    pdf = new_pdf()
    pdf_title(pdf, 'إقرار استلام عهدة', 'جمعية الإحسان للخدمات الاجتماعية')
    pdf_row(pdf, 'الموظف', employee.get('name'))
    pdf_row(pdf, 'المكتب / القسم', employee.get('office'))
    pdf_row(pdf, 'تاريخ الإقرار', date.today().isoformat())

    pdf_section(pdf, 'بيانات الجهاز المستلم')
    if device:
        pdf_row(pdf, 'اسم الجهاز', device.get('code'))
        pdf_row(pdf, 'النوع', device.get('type'))
        pdf_row(pdf, 'الرقم التسلسلي', device.get('serial'))
        pdf_row(pdf, 'المواصفات', f"{device.get('cpu') or ''} — {device.get('ram') or ''} — {device.get('storage') or ''}")
    else:
        pdf_row(pdf, 'الجهاز', 'ما فيه جهاز مرتبط حاليًا بهذا الموظف')

    pdf.ln(14)
    pdf.set_font('Body', '', 12)
    pdf.set_text_color(*PDF_TEXT)
    pdf_paragraph(pdf, 'أقر أنا الموظف المذكور أعلاه باستلام الجهاز الموضّحة بياناته، وأتحمّل المسؤولية الكاملة عن المحافظة عليه.')

    pdf.ln(20)
    y = pdf.get_y()
    pdf.set_draw_color(*PDF_LINE)
    pdf.line(18, y, 78, y)
    pdf.line(132, y, 192, y)
    pdf.set_font('Body', '', 11)
    pdf.set_text_color(*PDF_MUTED)
    pdf.set_xy(18, y + 2)
    pdf.cell(60, 7, rtl('توقيع الموظف'), align='R')
    pdf.set_xy(132, y + 2)
    pdf.cell(60, 7, rtl('توقيع مسؤول تقنية المعلومات'), align='R')

    return pdf_response(pdf, f"custody-{employee.get('name') or employee_id}.pdf")


@app.route('/api/disposal/<request_id>/pdf', methods=['GET'])
def disposal_pdf(request_id):
    err = require_page('disposal')
    if err:
        return err

    disposal_list = load_table('disposal_requests')
    req = next((r for r in disposal_list if r.get('id') == request_id), None)
    if not req:
        return jsonify({'error': 'الطلب غير موجود'}), 404

    snapshot = req.get('snapshot') or {}
    source_table = req.get('sourceTable')
    label = snapshot.get('code') or snapshot.get('name') or 'قطعة'

    pdf = new_pdf()
    pdf_title(pdf, 'طلب إتلاف أصل', 'جمعية الإحسان للخدمات الاجتماعية')
    pdf_row(pdf, 'نوع الأصل', DISPOSAL_SOURCE_LABELS.get(source_table, source_table))
    pdf_row(pdf, 'الحالة', req.get('status'))
    pdf_row(pdf, 'سبب الإتلاف', req.get('reason'))
    pdf_row(pdf, 'تاريخ الطلب', req.get('requestedAt'))
    if req.get('decidedAt'):
        pdf_row(pdf, 'تاريخ القرار', req.get('decidedAt'))

    pdf_section(pdf, 'تفاصيل القطعة وقت الطلب')
    schema = TABLE_SCHEMAS.get(source_table, [])
    if schema:
        row_values = record_to_row(source_table, snapshot)
        for (_key, col_label), value in zip(schema, row_values):
            pdf_row(pdf, col_label, value)

    pdf.ln(10)
    pdf.set_font('Body', '', 12)
    pdf.set_text_color(*PDF_TEXT)
    pdf_paragraph(pdf, 'نقر بأن القطعة الموضّحة بياناتها أعلاه غير قابلة للاستخدام أو لا حاجة لها، ونوافق على إتلافها/التصرف بها حسب الإجراءات المعتمدة.')

    pdf.ln(12)
    for sig_label in [
        'توقيع مدير وحدة التقنية والتحول الرقمي',
        'موافقة مدير الشؤون المالية والإدارية',
        'اعتماد المدير التنفيذي',
    ]:
        y = pdf.get_y()
        pdf.set_draw_color(*PDF_LINE)
        pdf.line(18, y, 192, y)
        pdf.set_font('Body', '', 11)
        pdf.set_text_color(*PDF_MUTED)
        pdf.set_xy(18, y + 2)
        pdf.cell(174, 7, rtl(sig_label), align='R')
        pdf.ln(18)

    return pdf_response(pdf, f"disposal-{label}.pdf")


# ------------------------------------------------------------------
# تصدير / استيراد Excel
# ------------------------------------------------------------------
@app.route('/api/<table>/export', methods=['GET'])
def export_table(table):
    if table not in TABLE_SCHEMAS:
        return jsonify({'error': 'جدول غير معروف'}), 404
    err = require_admin() if table in ADMIN_ONLY_TABLES else require_login()
    if err:
        return err

    records = load_table(table)
    schema = TABLE_SCHEMAS[table]

    wb = Workbook()
    ws = wb.active
    ws.title = table[:31]
    ws.append([label for _key, label in schema] + [ID_COLUMN_LABEL])
    for record in records:
        ws.append(record_to_row(table, record) + [record.get('id', '')])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True, download_name=f'{table}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/api/<table>/import', methods=['POST'])
def import_table(table):
    if table not in TABLE_SCHEMAS:
        return jsonify({'error': 'جدول غير معروف'}), 404
    err = require_admin() if table in ADMIN_ONLY_TABLES else require_login()
    if err:
        return err

    uploaded = request.files.get('file')
    if not uploaded:
        return jsonify({'error': 'لم يتم اختيار ملف'}), 400

    try:
        wb = load_workbook(uploaded, data_only=True)
        ws = wb.active
    except Exception:
        return jsonify({'error': 'تعذّر قراءة الملف — تأكد إنه ملف Excel صالح (xlsx)'}), 400

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({'error': 'الملف فاضي'}), 400
    headers = [str(h).strip() if h is not None else '' for h in rows[0]]

    records = load_table(table)
    by_id = {r['id']: r for r in records if r.get('id')}

    added = 0
    updated = 0
    skipped = []
    for raw_row in rows[1:]:
        if raw_row is None or all(c is None or str(c).strip() == '' for c in raw_row):
            continue
        row_dict = dict(zip(headers, raw_row))
        record_id = str(row_dict.get(ID_COLUMN_LABEL) or '').strip()
        existing = by_id.get(record_id) if record_id else None

        record = row_to_record(table, row_dict, existing)
        row_label = record.get('name') or record.get('code') or record.get('linked') or record_id or '؟'

        if table == 'employees' and record.get('device'):
            owner = next(
                (r for r in records if r.get('device') == record['device'] and r.get('id') != (existing or {}).get('id')),
                None,
            )
            if owner:
                skipped.append(f"{row_label}: الجهاز {record['device']} بعهدة {owner.get('name')} بالفعل")
                continue

        if existing:
            record['id'] = existing['id']
            idx = next(i for i, r in enumerate(records) if r['id'] == existing['id'])
            records[idx] = record
            updated += 1
        else:
            record['id'] = uuid.uuid4().hex
            records.append(record)
            by_id[record['id']] = record
            added += 1

    save_table(table, records)
    return jsonify({'success': True, 'added': added, 'updated': updated, 'skipped': skipped})


# ------------------------------------------------------------------
# تقديم واجهة الموقع
# ------------------------------------------------------------------
@app.route('/')
def index():
    return send_from_directory(STATIC_DIR, 'index.html')


if __name__ == '__main__':
    print('=' * 60)
    print('نظام الأصول والمخزون يعمل الآن.')
    print('افتح على نفس الجهاز: http://localhost:5000')
    print('شارك مع باقي الموظفين على نفس الشبكة عبر عنوان IP هذا الجهاز، مثال:')
    print('http://192.168.1.10:5000')
    print('=' * 60)
    app.run(host='0.0.0.0', port=5000, debug=False)
